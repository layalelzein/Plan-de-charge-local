import os
from flask import Flask, render_template, request, redirect, url_for, flash
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

app = Flask(__name__)
app.secret_key = 'secret_key'  # Nécessaire pour les messages flash

# Configurer le logging
logging.basicConfig(level=logging.DEBUG)

# Définir le chemin absolu vers le fichier data.xlsx
base_dir = os.path.dirname(os.path.abspath(__file__))
data_file_path = os.path.join(base_dir, 'data.xlsx')


# Charger les données existantes
def load_data():
    global df_clients, df_agents, df_affaires, df_plan_charge, df_dates
    df_clients = pd.read_excel(data_file_path, sheet_name='Clients', engine='openpyxl')
    df_agents = pd.read_excel(data_file_path, sheet_name='Agents', engine='openpyxl')
    df_affaires = pd.read_excel(data_file_path, sheet_name='Affaires', engine='openpyxl')
    df_plan_charge = pd.read_excel(data_file_path, sheet_name='Plan de Charge', engine='openpyxl')
    df_dates = pd.read_excel(data_file_path, sheet_name='Dates', engine='openpyxl')

    # Convertir les colonnes appropriées en chaînes de caractères
    df_clients['ClientID'] = df_clients['ClientID'].astype(str).str.strip()
    df_plan_charge['ClientID'] = df_plan_charge['ClientID'].astype(str).str.strip()

    df_agents['AgentID'] = df_agents['AgentID'].astype(str)
    df_affaires['NumeroAffaire'] = df_affaires['NumeroAffaire'].astype(str)
    df_affaires['AffaireID'] = df_affaires['AffaireID'].astype(str)
    df_plan_charge['PlanChargeID'] = df_plan_charge['PlanChargeID'].astype(str)
    df_plan_charge['AffaireID'] = df_plan_charge['AffaireID'].astype(str)
    df_plan_charge['AgentID'] = df_plan_charge['AgentID'].astype(str)

    # Vérifier et corriger les dates
    df_plan_charge['Date'] = pd.to_datetime(df_plan_charge['Date'], errors='coerce')

    # Filtrer les dates non convertibles et afficher des avertissements
    invalid_dates = df_plan_charge[df_plan_charge['Date'].isna()]
    if not invalid_dates.empty:
        logging.warning(f"Les dates suivantes sont invalides et ont été ignorées :\n{invalid_dates[['Date']]}")

    # Supprimer les lignes avec des dates non valides
    df_plan_charge = df_plan_charge.dropna(subset=['Date'])

    logging.debug("Data loaded successfully")
    logging.debug(f"df_affaires: {df_affaires.head()}")
    logging.debug(f"df_plan_charge: {df_plan_charge.head()}")

load_data()

# Statuts et étapes prédéfinis
STATUTS = ['Terminé', 'Abandonnée', 'Non planifiée', 'En attente']
ENTITES = ['E. Bat', 'E. Ist', 'T. Bat', 'T. Ist']
ETAPES = ['Consultation', 'Etude', 'Travaux']

@app.route('/')
def index():
    clients = df_clients.to_dict('records')
    agents = df_agents.to_dict('records')
    affaires = df_affaires.to_dict('records')
    return render_template('index.html', clients=clients, agents=agents, affaires=affaires, statuts=STATUTS, entites=ENTITES, etapes=ETAPES)


@app.route('/submit', methods=['POST'])
def submit():
    logging.debug("Submit function called")
    global df_clients, df_agents, df_affaires, df_plan_charge, df_dates

    client_id = request.form.get('ClientID', '').strip()
    statut = request.form.get('Statut', '').strip()
    numero_affaire = request.form.get('NumeroAffaire', '').strip()
    new_numero_affaire = request.form.get('NewNumeroAffaire', '').strip()
    intitule = request.form.get('IntituleSelect', '').strip() or request.form.get('Intitule', '').strip()
    new_intitule = request.form.get('NewIntitule', '').strip()
    agent_id = request.form.get('AgentID', '').strip()
    new_agent_name = request.form.get('NewAgentName', '').strip()
    montant = request.form.get('Montant', '').strip()
    entite = request.form.get('Entite', '').strip()
    date_debut = request.form.get('DateDebut', '').strip()
    date_fin = request.form.get('DateFin', '').strip()
    pourcentage_charge = request.form.get('PourcentageCharge', '').strip()
    etape = request.form.get('Etape', '').strip()

    logging.debug(f"ClientID: {client_id}, Statut: {statut}, NumeroAffaire: {numero_affaire}, NewNumeroAffaire: {new_numero_affaire}, Intitule: {intitule}, NewIntitule: {new_intitule}, AgentID: {agent_id}, NewAgentName: {new_agent_name}, Montant: {montant}, Entite: {entite}, DateDebut: {date_debut}, DateFin: {date_fin}, PourcentageCharge: {pourcentage_charge}, Etape: {etape}")

    # Vérifier les champs obligatoires
    if not (numero_affaire or new_numero_affaire):
        flash("Erreur: Veuillez fournir un numéro d'affaire ou en créer un nouveau.", "error")
        return redirect(url_for('index'))

    if not (intitule or new_intitule):
        flash("Erreur: Veuillez fournir un intitulé ou en créer un nouveau.", "error")
        return redirect(url_for('index'))

    if agent_id == '--Nouveau agent--' and not new_agent_name:
        flash("Erreur: Veuillez fournir le nom du nouveau agent.", "error")
        return redirect(url_for('index'))
    elif agent_id == '' and not new_agent_name:
        flash("Erreur: Veuillez sélectionner un agent ou en créer un nouveau.", "error")
        return redirect(url_for('index'))

    if not date_debut:
        flash("Erreur: Veuillez fournir une date de début.", "error")
        return redirect(url_for('index'))

    if not date_fin:
        flash("Erreur: Veuillez fournir une date de fin.", "error")
        return redirect(url_for('index'))

    if not pourcentage_charge:
        flash("Erreur: Veuillez fournir un pourcentage de charge.", "error")
        return redirect(url_for('index'))

    # Convertir le pourcentage de charge en entier
    try:
        pourcentage_charge = int(float(pourcentage_charge))
    except ValueError:
        flash("Erreur: Le pourcentage de charge doit être un nombre.", "error")
        return redirect(url_for('index'))

    # Utiliser le nouveau numéro d'affaire s'il est fourni
    if new_numero_affaire:
        logging.debug(f"Nouveau numéro d'affaire: {new_numero_affaire}")
        numero_affaire = new_numero_affaire
        if not new_intitule:
            flash("Erreur: Un nouvel intitulé doit être fourni avec un nouveau numéro d'affaire.", "error")
            return redirect(url_for('index'))
        intitule = new_intitule

        # Créer une nouvelle affaire
        affaire_id = str(int(df_affaires['AffaireID'].astype(int).max()) + 1 if not df_affaires.empty else 1)
        new_affaire = pd.DataFrame([{
            'AffaireID': affaire_id,
            'ClientID': client_id,
            'NumeroAffaire': numero_affaire,
            'Intitule': intitule
        }])
        df_affaires = pd.concat([df_affaires, new_affaire], ignore_index=True)
        df_affaires.to_excel(data_file_path, sheet_name='Affaires', index=False, engine='openpyxl')
        logging.debug(f"Nouvelle affaire créée avec AffaireID: {affaire_id}")
    else:
        # Utiliser l'affaire existante
        logging.debug(f"Recherche d'affaire existante pour NumeroAffaire: {numero_affaire} de type {type(numero_affaire)}")
        df_affaires['NumeroAffaire'] = df_affaires['NumeroAffaire'].astype(str)
        existing_affaire = df_affaires[df_affaires['NumeroAffaire'] == numero_affaire]

        logging.debug(f"Résultat de la recherche: {existing_affaire}")

        if not existing_affaire.empty:
            affaire_id = existing_affaire.iloc[0]['AffaireID']
            intitule = existing_affaire.iloc[0]['Intitule']
            logging.debug(f"Affaire existante trouvée: {affaire_id} - {intitule}")
        else:
            flash("Erreur: L'affaire n'existe pas et aucun nouveau numéro d'affaire n'a été fourni.", "error")
            return redirect(url_for('index'))

    # Vérifier si l'agent existe, sinon créer un nouveau
    if agent_id == '--Nouveau agent--':
        agent_id = str(int(df_agents['AgentID'].astype(int).max()) + 1 if not df_agents.empty else 1)
        new_agent = pd.DataFrame([{
            'AgentID': agent_id,
            'AgentName': new_agent_name
        }])
        df_agents = pd.concat([df_agents, new_agent], ignore_index=True)
        df_agents.to_excel(data_file_path, sheet_name='Agents', index=False, engine='openpyxl')
        logging.debug(f"Nouvel agent créé avec AgentID: {agent_id}")
    elif agent_id != '':
        agent_id = str(agent_id)

    # Convertir les dates en datetime pour la comparaison
    date_debut_dt = datetime.strptime(date_debut, '%Y-%m-%d')
    date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d')

    # Vérifier les chevauchements de périodes
    existing_entries = df_plan_charge[
        (df_plan_charge['AffaireID'] == affaire_id) & 
        (df_plan_charge['AgentID'] == agent_id) & 
        (pd.to_datetime(df_plan_charge['Date']) >= date_debut_dt) &
        (pd.to_datetime(df_plan_charge['Date']) <= date_fin_dt)
    ]
    if not existing_entries.empty:
        overlap_dates = pd.to_datetime(existing_entries['Date']).dt.strftime('%Y-%m-%d').tolist()
        overlap_dates_str = ', '.join(overlap_dates)
        flash(f"Erreur: Les dates suivantes sont déjà chargées pour cet agent sur cette affaire: {overlap_dates_str}", "error")
        return redirect(url_for('index'))

    # Calculer le nouvel ID pour Plan de Charge
    if not df_plan_charge.empty:
        max_plan_charge_id = int(df_plan_charge['PlanChargeID'].astype(int).max())
        logging.debug(f"PlanChargeID maximum actuel: {max_plan_charge_id}")
        plan_charge_id = str(max_plan_charge_id + 1)
    else:
        plan_charge_id = '1'
    logging.debug(f"Nouveau PlanChargeID: {plan_charge_id}")

    # Ajouter les nouvelles entrées de charge
    current_date = date_debut_dt
    total_charge_alerts = []

    while current_date <= date_fin_dt:
        data = pd.DataFrame([{
            "PlanChargeID": plan_charge_id,
            "ClientID": client_id,
            "Statut": statut,
            "AffaireID": affaire_id,
            "AgentID": agent_id,
            "Montant": montant,
            "Entite": entite,
            "Date": current_date.strftime('%Y-%m-%d'),
            "PourcentageCharge": pourcentage_charge,
            "Etape": etape
        }])
        
        df_plan_charge = pd.concat([df_plan_charge, data], ignore_index=True)
        
        # Vérifier le total des charges pour cette semaine
        week_start = current_date - timedelta(days=current_date.weekday())
        week_end = week_start + timedelta(days=6)
        week_entries = df_plan_charge[
            (df_plan_charge['AgentID'] == agent_id) &
            (pd.to_datetime(df_plan_charge['Date']) >= week_start) &
            (pd.to_datetime(df_plan_charge['Date']) <= week_end)
        ]
        total_charge = week_entries['PourcentageCharge'].sum()

        if total_charge > 100:
            total_charge_alerts.append(current_date.strftime('%Y-%m-%d'))

        logging.debug(f"Ajout de données pour la date {current_date.strftime('%Y-%m-%d')} avec PlanChargeID: {plan_charge_id}")
        current_date += timedelta(weeks=1)
        plan_charge_id = str(int(plan_charge_id) + 1)
        logging.debug(f"PlanChargeID incrémenté: {plan_charge_id}")

    # Sauvegarder toutes les feuilles dans le fichier Excel
    logging.debug("Enregistrement des données dans le fichier Excel...")
    try:
        with pd.ExcelWriter(data_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_clients.to_excel(writer, sheet_name='Clients', index=False)
            df_agents.to_excel(writer, sheet_name='Agents', index=False)
            df_affaires.to_excel(writer, sheet_name='Affaires', index=False)
            df_plan_charge.to_excel(writer, sheet_name='Plan de Charge', index=False)
            df_dates.to_excel(writer, sheet_name='Dates', index=False)
        logging.debug("Données enregistrées avec succès.")
    except Exception as e:
        logging.error(f"Erreur lors de l'enregistrement des données : {e}")
        raise

    load_data()  # Recharger les données pour assurer la cohérence

    if total_charge_alerts:
        total_charge_alerts_str = ', '.join(total_charge_alerts)
        flash(f"Alerte: Les semaines suivantes dépassent 100% de charge pour cet agent: {total_charge_alerts_str}", "warning")

    # Appeler la fonction pour mettre à jour la feuille "Visu"
    create_visu_sheet(data_file_path)

    flash("Enregistrement réussi !", "success")
    return redirect(url_for('index'))


@app.route('/edit/<plan_charge_id>', methods=['GET', 'POST'])
def edit(plan_charge_id):
    load_data()  # Recharger les données pour assurer la cohérence

    # Convertir les ID en chaînes pour éviter les problèmes de compatibilité
    df_clients['ClientID'] = df_clients['ClientID'].astype(str)
    df_affaires['ClientID'] = df_affaires['ClientID'].astype(str)
    df_plan_charge['ClientID'] = df_plan_charge['ClientID'].astype(str)
    df_plan_charge['AffaireID'] = df_plan_charge['AffaireID'].astype(str)

    # Trouver l'enregistrement correspondant au plan_charge_id
    entry = df_plan_charge[df_plan_charge['PlanChargeID'] == plan_charge_id]

    if entry.empty:
        flash("Erreur: L'enregistrement demandé n'existe pas.", "error")
        return redirect(url_for('records'))

    entry = entry.iloc[0]  # Sélectionner la première entrée correspondante

    # Assurez-vous que la date est au format string pour éviter les problèmes
    if not pd.isnull(entry['Date']):
        entry['Date'] = entry['Date'].strftime('%Y-%m-%d')

    # Remplacer les NaN par des chaînes vides pour l'affichage
    entry = entry.fillna('')

    # Convertir le pourcentage de charge en entier pour l'affichage
    if 'PourcentageCharge' in entry:
        entry['PourcentageCharge'] = int(float(entry['PourcentageCharge']))

    if request.method == 'POST':
        try:
            # Récupérer les valeurs depuis le formulaire, sauf `NumeroAffaire`, `Intitule` et `Date`
            client_id = request.form.get('ClientID', '').strip()
            statut = request.form.get('Statut', '').strip()
            agent_id = request.form.get('AgentID', '').strip()
            montant = request.form.get('Montant', '').strip()
            entite = request.form.get('Entite', '').strip()
            pourcentage_charge = request.form.get('PourcentageCharge', '').strip()
            etape = request.form.get('Etape', '').strip()

            # Convertir les types appropriés
            if montant:
                montant = float(montant)
            if pourcentage_charge:
                pourcentage_charge = int(float(pourcentage_charge))

            # Mettre à jour l'enregistrement, sauf `NumeroAffaire`, `Intitule` et `Date`
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'ClientID'] = client_id
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'Statut'] = statut
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'AgentID'] = agent_id
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'Montant'] = montant
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'Entite'] = entite
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'PourcentageCharge'] = pourcentage_charge
            df_plan_charge.loc[df_plan_charge['PlanChargeID'] == plan_charge_id, 'Etape'] = etape

            # Sauvegarder les modifications
            with pd.ExcelWriter(data_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_clients.to_excel(writer, sheet_name='Clients', index=False)
                df_agents.to_excel(writer, sheet_name='Agents', index=False)
                df_affaires.to_excel(writer, sheet_name='Affaires', index=False)
                df_plan_charge.to_excel(writer, sheet_name='Plan de Charge', index=False, columns=[
                    'PlanChargeID', 'ClientID', 'Statut', 'AffaireID', 'AgentID', 'Montant', 'Entite', 'Date', 'PourcentageCharge', 'Etape'
                ])
                df_dates.to_excel(writer, sheet_name='Dates', index=False)

            # Mettre à jour la feuille Visu après l'édition
            create_visu_sheet(data_file_path)
            
            flash("Enregistrement mis à jour avec succès!", "success")
            return redirect(url_for('records')) 

        except Exception as e:
            logging.error(f"Erreur lors de la mise à jour : {e}")
            flash(f"Erreur lors de la mise à jour : {e}", "error")


    # Vérification et récupération des informations sur l'affaire
    affaire = df_affaires[df_affaires['AffaireID'] == entry['AffaireID']]
    if not affaire.empty:
        affaire = affaire.iloc[0]

        client_info = df_clients[df_clients['ClientID'] == affaire['ClientID']]
        if not client_info.empty:
            entry['ClientName'] = client_info.iloc[0]['ClientName']
        else:
            entry['ClientName'] = 'Inconnu'

        entry['NumeroAffaire'] = affaire['NumeroAffaire']
        entry['Intitule'] = affaire['Intitule']
    else:
        entry['ClientName'] = 'Inconnu'
        entry['NumeroAffaire'] = 'Inconnu'
        entry['Intitule'] = 'Inconnu'

    # Récupérer les informations nécessaires pour le formulaire
    clients = df_clients.to_dict('records')
    agents = df_agents.to_dict('records')
    affaires = df_affaires.to_dict('records')
    statuts = STATUTS
    entites = ENTITES
    etapes = ETAPES

    return render_template('edit.html', entry=entry, clients=clients, agents=agents, affaires=affaires, statuts=statuts, entites=entites, etapes=etapes)


@app.route('/delete/<plan_charge_id>', methods=['GET'])
def delete(plan_charge_id):
    logging.debug(f"Delete function called for PlanChargeID: {plan_charge_id}")

    global df_plan_charge
    df_plan_charge = df_plan_charge[df_plan_charge['PlanChargeID'] != plan_charge_id]

    # Sauvegarder les modifications
    with pd.ExcelWriter(data_file_path, engine='openpyxl') as writer:
        df_clients.to_excel(writer, sheet_name='Clients', index=False)
        df_agents.to_excel(writer, sheet_name='Agents', index=False)
        df_affaires.to_excel(writer, sheet_name='Affaires', index=False)
        df_plan_charge.to_excel(writer, sheet_name='Plan de Charge', index=False)
        df_dates.to_excel(writer, sheet_name='Dates', index=False)

    # Mettre à jour la feuille Visu après la suppression
    create_visu_sheet(data_file_path)
    
    flash("Suppression réussie!", "success")
    return redirect(url_for('records'))


@app.route('/records')
def records():
    # Charger les données existantes pour les afficher
    load_data()

    # Convertir les colonnes en chaînes de caractères pour assurer la compatibilité
    df_plan_charge['ClientID'] = df_plan_charge['ClientID'].astype(str)
    df_plan_charge['AffaireID'] = df_plan_charge['AffaireID'].astype(str)
    df_plan_charge['AgentID'] = df_plan_charge['AgentID'].astype(str)

    df_clients['ClientID'] = df_clients['ClientID'].astype(str)
    df_affaires['AffaireID'] = df_affaires['AffaireID'].astype(str)
    df_agents['AgentID'] = df_agents['AgentID'].astype(str)

    # Fusionner les DataFrames pour obtenir les noms correspondants
    df_records = df_plan_charge.merge(df_clients, on='ClientID', how='left') \
                               .merge(df_affaires, on='AffaireID', how='left') \
                               .merge(df_agents, on='AgentID', how='left')

    # Convertir les pourcentages de charge en entiers pour l'affichage
    df_records['PourcentageCharge'] = df_records['PourcentageCharge'].astype(int)

    # Remplacer les NaN par des chaînes vides pour l'affichage
    df_records = df_records.fillna('')

    # Formater les dates en JJ/MM/AAAA
    df_records['Date'] = df_records['Date'].dt.strftime('%d/%m/%Y')

    entries = df_records.to_dict('records')
    return render_template('records.html', entries=entries)


def create_visu_sheet(data_file_path):
    # Fonction pour convertir l'index de colonne en lettre
    def column_index_to_letter(column_index):
        column_letter = ''
        while column_index > 0:
            column_index, remainder = divmod(column_index - 1, 26)
            column_letter = chr(65 + remainder) + column_letter
        return column_letter

    # Charger les données
    df_clients = pd.read_excel(data_file_path, sheet_name='Clients')
    df_agents = pd.read_excel(data_file_path, sheet_name='Agents')
    df_affaires = pd.read_excel(data_file_path, sheet_name='Affaires')
    df_plan_charge = pd.read_excel(data_file_path, sheet_name='Plan de Charge')

    # Assurer que les types sont bien des chaînes
    df_clients['ClientID'] = df_clients['ClientID'].astype(str).str.strip()
    df_plan_charge['ClientID'] = df_plan_charge['ClientID'].astype(str).str.strip()

    # Assurer que la colonne Date est en datetime
    df_plan_charge['Date'] = pd.to_datetime(df_plan_charge['Date'], errors='coerce')
    df_plan_charge = df_plan_charge.dropna(subset=['Date'])

    # Charger le classeur
    wb = load_workbook(data_file_path)
    
    # Supprimer la feuille "Visu" si elle existe
    if "Visu" in wb.sheetnames:
        del wb["Visu"]

    # Ajouter une nouvelle feuille "Visu"
    ws_visu = wb.create_sheet(title="Visu")

    # Ajouter les en-têtes
    headers = ["ClientName", "Statut", "NumeroAffaire", "Intitule", "AgentName", "Montant", "Entite"]
    ws_visu.append(headers)

    # Ajouter les colonnes des semaines à partir du 01/01/2024
    current_date = pd.Timestamp('2024-01-01')
    max_date = df_plan_charge['Date'].max()
    date_columns = {}  # Dictionnaire pour stocker la correspondance des dates et colonnes
    while current_date <= max_date:
        col_index = len(headers) + 1 + ((current_date - pd.Timestamp('2024-01-01')).days // 7)
        ws_visu.cell(row=1, column=col_index, value=current_date.strftime('%d/%m/%Y'))
        date_columns[current_date] = col_index
        current_date += pd.Timedelta(weeks=1)

    # Initialiser last_col avec une valeur par défaut
    last_col = len(headers)

    # Utiliser un dictionnaire pour stocker les lignes existantes
    existing_rows = {}

    # Collecter toutes les lignes de données
    all_rows = []

    # Remplir les données
    for _, plan_row in df_plan_charge.iterrows():
        # Vérifier que le ClientID existe dans df_clients
        client_row = df_clients.loc[df_clients['ClientID'] == plan_row['ClientID']]
        if not client_row.empty:
            client_name = client_row['ClientName'].values[0]
        else:
            client_name = 'Client inconnu'
            print(f"Erreur: ClientID {plan_row['ClientID']} n'existe pas dans df_clients")

        # Vérifier que l'AgentID existe dans df_agents
        agent_row = df_agents.loc[df_agents['AgentID'] == plan_row['AgentID']]
        if not agent_row.empty:
            agent_name = agent_row['AgentName'].values[0]
        else:
            agent_name = 'Agent inconnu'
            print(f"Erreur: AgentID {plan_row['AgentID']} n'existe pas dans df_agents")

        # Vérifier que l'AffaireID existe dans df_affaires
        affaire_row = df_affaires.loc[df_affaires['AffaireID'] == plan_row['AffaireID']]
        if not affaire_row.empty:
            numero_affaire = affaire_row['NumeroAffaire'].values[0]
            intitule = affaire_row['Intitule'].values[0]
        else:
            numero_affaire = 'Affaire inconnue'
            intitule = 'Affaire inconnue'
            print(f"Erreur: AffaireID {plan_row['AffaireID']} n'existe pas dans df_affaires")

        # Créer une clé unique pour la ligne
        unique_key = (client_name, numero_affaire, intitule, agent_name, plan_row['Entite'])

        # Ajouter les données à la liste all_rows
        row_data = [client_name, plan_row['Statut'], numero_affaire, intitule, agent_name, plan_row['Montant'], plan_row['Entite']]
        for date in date_columns:
            col_index = date_columns[date]
            if plan_row['Date'] == date:
                row_data.append((col_index, plan_row['PourcentageCharge'], plan_row['Etape']))
        all_rows.append((unique_key, row_data))

    # Trier les données par colonne Intitule
    all_rows.sort(key=lambda x: x[1][3])

    # Insérer les données triées dans la feuille
    for unique_key, row_data in all_rows:
        if unique_key in existing_rows:
            row = existing_rows[unique_key]
        else:
            row = ws_visu.max_row + 1
            existing_rows[unique_key] = row
            ws_visu.append(row_data[:7])

        # Remplir les pourcentages de charge pour les dates correspondantes
        for col_index, pourcentage, etape in row_data[7:]:
            ws_visu.cell(row=row, column=col_index, value=pourcentage)

            # Appliquer la couleur en fonction de l'étape
            etape_color_map = {
                "Consultation": "FFFF99",  # Jaune pâle
                "Etude": "FFC0CB",        # Rose pâle
                "Travaux": "CCFF99"       # Vert pâle
            }
            etape_color = etape_color_map.get(etape, "FFFFFF")
            fill = PatternFill(start_color=etape_color, end_color=etape_color, fill_type="solid")
            ws_visu.cell(row=row, column=col_index).fill = fill

        # Mettre à jour last_col pour s'assurer qu'il contient l'indice de la dernière colonne utilisée
        last_col = max(last_col, col_index)

    # Calculer les totaux par agent et par semaine
    agent_totals = df_plan_charge.groupby(['AgentID', pd.Grouper(key='Date', freq='W-MON')])['PourcentageCharge'].sum().reset_index()

    # Utiliser un dictionnaire pour éviter les doublons
    agent_rows = {}

    # Ajouter les totaux dans la feuille "Visu"
    for _, total_row in agent_totals.iterrows():
        agent_row = df_agents.loc[df_agents['AgentID'] == total_row['AgentID']]
        if not agent_row.empty:
            agent_name = agent_row['AgentName'].values[0]
        else:
            agent_name = 'Agent inconnu'
            print(f"Erreur: AgentID {total_row['AgentID']} n'existe pas dans df_agents")

        # Ajouter ou mettre à jour une ligne pour les totaux de l'agent
        if agent_name in agent_rows:
            row = agent_rows[agent_name]
        else:
            row = ws_visu.max_row + 1
            agent_rows[agent_name] = row
            ws_visu.cell(row=row, column=5, value=agent_name)  # Colonne AgentName

        week_start = total_row['Date'] - timedelta(days=total_row['Date'].weekday())
        col_index = date_columns.get(week_start)

        # Remplir les données du total
        current_value = ws_visu.cell(row=row, column=col_index).value
        if current_value is None:
            ws_visu.cell(row=row, column=col_index, value=total_row['PourcentageCharge'])
        else:
            ws_visu.cell(row=row, column=col_index, value=current_value + total_row['PourcentageCharge'])

        # Appliquer un style distinct pour les lignes de totaux
        for col in range(1, last_col + 1):
            cell = ws_visu.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    # Définir la plage de cellules pour le tableau
    last_row = ws_visu.max_row
    table_range = f"A1:{column_index_to_letter(last_col)}{last_row}"

    # Créer le tableau
    table = Table(displayName="VisuTable", ref=table_range)

    # Ajouter le style au tableau
    style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Ajouter le tableau à la feuille de calcul
    ws_visu.add_table(table)

    # Enregistrer les modifications
    wb.save(data_file_path)


if __name__ == '__main__':
    app.run(debug=True, port=5002)
